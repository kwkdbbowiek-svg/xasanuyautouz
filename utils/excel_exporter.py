"""
Export all users and their subscription/ad data to an Excel file.
Uses openpyxl for clean formatting.
"""
from __future__ import annotations

import os
import tempfile
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select

from database.connection import AsyncSessionFactory
from database.models import Ad, AdStatus, Subscription, User, UserRole


ROLE_LABELS = {
    UserRole.seller: "Sotuvchi",
    UserRole.buyer: "Oluvchi",
    UserRole.owner: "Kvartira egasi",
    UserRole.seeker: "Kvartira qidiruvchi",
    None: "Belgilanmagan",
}

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL = PatternFill(start_color="D6E4F7", end_color="D6E4F7", fill_type="solid")


async def export_users_to_excel() -> str:
    """
    Create an Excel workbook with two sheets:
      1. Users — basic info, role, subscription status
      2. Ads — all ads with status
    Returns the temp file path (caller must delete it).
    """
    now = datetime.now(timezone.utc)

    async with AsyncSessionFactory() as session:
        # Fetch users
        users_result = await session.execute(select(User).order_by(User.created_at))
        users: list[User] = users_result.scalars().all()

        # Fetch active subscriptions per user
        subs_result = await session.execute(
            select(Subscription).where(
                Subscription.is_active == True,  # noqa: E712
                Subscription.expires_at > now,
            )
        )
        subs = subs_result.scalars().all()
        sub_map: dict[int, Subscription] = {s.user_id: s for s in subs}

        # Fetch ad counts per user
        ad_count_result = await session.execute(
            select(Ad.owner_id, func.count(Ad.id)).group_by(Ad.owner_id)
        )
        total_ad_map: dict[int, int] = {row[0]: row[1] for row in ad_count_result.all()}

        active_ad_result = await session.execute(
            select(Ad.owner_id, func.count(Ad.id))
            .where(Ad.status == AdStatus.active)
            .group_by(Ad.owner_id)
        )
        active_ad_map: dict[int, int] = {row[0]: row[1] for row in active_ad_result.all()}

        # Fetch all ads
        ads_result = await session.execute(select(Ad).order_by(Ad.created_at.desc()))
        ads: list[Ad] = ads_result.scalars().all()

    wb = Workbook()

    # ── Sheet 1: Users ────────────────────────────────────────────────────────
    ws_users = wb.active
    ws_users.title = "Foydalanuvchilar"

    user_headers = [
        "№", "Telegram ID", "To'liq ism", "Username", "Rol",
        "Faol obuna", "Obuna turi", "Obuna tugaydi",
        "Jami e'lonlar", "Faol e'lonlar", "Ro'yxatdan o'tgan sana",
    ]
    _write_headers(ws_users, user_headers)

    for idx, user in enumerate(users, start=1):
        sub = sub_map.get(user.id)
        sub_active = "Ha" if sub else "Yo'q"
        sub_type = sub.sub_type.value if sub else "—"
        sub_expires = sub.expires_at.strftime("%d.%m.%Y") if sub else "—"

        row = [
            idx,
            user.id,
            user.full_name,
            f"@{user.username}" if user.username else "—",
            ROLE_LABELS.get(user.role, "—"),
            sub_active,
            sub_type,
            sub_expires,
            total_ad_map.get(user.id, 0),
            active_ad_map.get(user.id, 0),
            user.created_at.strftime("%d.%m.%Y %H:%M") if user.created_at else "—",
        ]
        ws_row = ws_users.max_row + 1
        for col, value in enumerate(row, start=1):
            cell = ws_users.cell(row=ws_row, column=col, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if idx % 2 == 0:
                cell.fill = ALT_FILL

    _auto_width(ws_users)

    # ── Sheet 2: Ads ──────────────────────────────────────────────────────────
    ws_ads = wb.create_sheet("E'lonlar")

    ad_headers = [
        "№", "ID", "Egasi (Telegram ID)", "Blok ID", "Tur",
        "Obuna turi", "Sarlavha", "Narx (so'm)", "Holat",
        "Yaratilgan", "Tugaydi",
    ]
    _write_headers(ws_ads, ad_headers)

    for idx, ad in enumerate(ads, start=1):
        row = [
            idx,
            ad.id,
            ad.owner_id,
            ad.block_id,
            "Sotish" if ad.ad_type.value == "sale" else "Ijara",
            ad.sub_type.value,
            ad.title,
            ad.price,
            ad.status.value,
            ad.created_at.strftime("%d.%m.%Y %H:%M") if ad.created_at else "—",
            ad.expires_at.strftime("%d.%m.%Y") if ad.expires_at else "—",
        ]
        ws_row = ws_ads.max_row + 1
        for col, value in enumerate(row, start=1):
            cell = ws_ads.cell(row=ws_row, column=col, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if idx % 2 == 0:
                cell.fill = ALT_FILL

    _auto_width(ws_ads)

    # Save to temp file
    tmp = tempfile.NamedTemporaryFile(
        suffix=".xlsx", delete=False, prefix="uysavdo_export_"
    )
    tmp.close()
    wb.save(tmp.name)
    return tmp.name


def _write_headers(ws, headers: list[str]) -> None:
    ws.append(headers)
    for col, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value else 0
                if cell_len > max_len:
                    max_len = cell_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)
